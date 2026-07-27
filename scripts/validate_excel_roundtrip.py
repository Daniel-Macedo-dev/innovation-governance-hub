from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from innovation_governance_hub.database import Base
from innovation_governance_hub.excel.exporters import executive_workbook
from innovation_governance_hub.excel.importers import (
    persist_preview,
    preview_ai_cases,
    preview_expenses,
    preview_indicators,
    preview_initiatives,
)
from innovation_governance_hub.excel.templates import (
    AI_CASE_COLUMNS,
    EXPENSE_COLUMNS,
    INDICATOR_COLUMNS,
    INITIATIVE_COLUMNS,
    TEMPLATE_FILES,
    create_template,
)
from innovation_governance_hub.persistence.models import (
    AIUseCase,
    Expense,
    Initiative,
    InitiativeIndicator,
)


def workbook_bytes(columns: list[str], row: dict[str, object]) -> bytes:
    output = BytesIO()
    pd.DataFrame([row], columns=columns).to_excel(output, index=False)
    return output.getvalue()


def main() -> None:
    for filename, columns in TEMPLATE_FILES.values():
        create_template(Path("templates") / filename, columns)
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session, session.begin():
        initiative_row = dict(
            zip(
                INITIATIVE_COLUMNS,
                [
                    "INI-RT-001",
                    "Iniciativa roundtrip",
                    "Problema fictício",
                    "Solução fictícia",
                    "Operações",
                    "Pessoa Demo",
                    "Alta",
                    "Alto",
                    "Impacto fictício",
                    "Média",
                    date(2026, 7, 1),
                    date(2026, 12, 1),
                    "Ativa",
                    "Ideia",
                    10000,
                    20000,
                    "Validação isolada",
                ],
                strict=True,
            )
        )
        initiatives = preview_initiatives(workbook_bytes(INITIATIVE_COLUMNS, initiative_row))
        assert initiatives.valid
        persist_preview(session, initiatives)
        item = session.scalar(select(Initiative).where(Initiative.code == "INI-RT-001"))
        assert item is not None
        expense_row = dict(
            zip(
                EXPENSE_COLUMNS,
                [
                    date(2026, 7, 1),
                    item.code,
                    "Consultoria",
                    "Despesa roundtrip",
                    "Fornecedor fictício",
                    "",
                    "Pontual",
                    "Realizado",
                    500,
                ],
                strict=True,
            )
        )
        expenses = preview_expenses(
            workbook_bytes(EXPENSE_COLUMNS, expense_row), {item.code: item.id}
        )
        assert expenses.valid
        persist_preview(session, expenses)
        assert session.scalar(select(func.count()).select_from(Expense)) == 1
        ai_row = dict(
            zip(
                AI_CASE_COLUMNS,
                [
                    "IA-RT-001",
                    "Caso roundtrip",
                    "Operações",
                    "Apoiar análise fictícia",
                    "Assistente fictício",
                    "Provedor avaliado",
                    "Dados sintéticos",
                    "Não",
                    "Médio",
                    "Revisão periódica",
                    "Ganho potencial",
                    "Em avaliação",
                    "Responsável Demo",
                    date(2026, 12, 1),
                    "Não",
                    "Não",
                    100,
                    25,
                    "Roundtrip",
                ],
                strict=True,
            )
        )
        ai_cases = preview_ai_cases(workbook_bytes(AI_CASE_COLUMNS, ai_row))
        assert ai_cases.valid, ai_cases.issues
        persist_preview(session, ai_cases)
        assert session.scalar(select(func.count()).select_from(AIUseCase)) == 1
        indicator_row = dict(
            zip(
                INDICATOR_COLUMNS,
                [
                    item.code,
                    "Tempo de ciclo roundtrip",
                    "Indicador fictício",
                    "Dias",
                    20,
                    8,
                    15,
                    "Reduzir",
                    "Responsável Demo",
                    date(2026, 7, 15),
                    "Roundtrip",
                ],
                strict=True,
            )
        )
        indicators = preview_indicators(
            workbook_bytes(INDICATOR_COLUMNS, indicator_row), {item.code: item.id}
        )
        assert indicators.valid, indicators.issues
        persist_preview(session, indicators)
        assert session.scalar(select(func.count()).select_from(InitiativeIndicator)) == 1
        data = executive_workbook(session)
    workbook = load_workbook(BytesIO(data))
    required = {
        "Resumo executivo",
        "Iniciativas",
        "Governança de IA",
        "Custos",
        "Pendências",
        "Definições das métricas",
    }
    if not required <= set(workbook.sheetnames):
        raise RuntimeError("Abas obrigatórias ausentes")
    Path("templates/relatorio_executivo_validacao.xlsx").write_bytes(data)
    print("Roundtrip Excel validado.")


if __name__ == "__main__":
    main()
