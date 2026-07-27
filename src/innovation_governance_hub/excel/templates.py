from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

INITIATIVE_COLUMNS = [
    "Código",
    "Nome",
    "Descrição do problema",
    "Solução proposta",
    "Área solicitante",
    "Responsável",
    "Prioridade",
    "Impacto esperado",
    "Descrição do impacto",
    "Complexidade",
    "Data de criação",
    "Prazo",
    "Status",
    "Estágio atual",
    "Custo planejado",
    "Benefício esperado",
    "Observações",
]
EXPENSE_COLUMNS = [
    "Data de competência",
    "Código da iniciativa",
    "Categoria",
    "Descrição",
    "Fornecedor",
    "Ferramenta",
    "Tipo de custo",
    "Status financeiro",
    "Valor",
]
AI_CASE_COLUMNS = [
    "Código",
    "Nome",
    "Área responsável",
    "Objetivo",
    "Ferramenta avaliada",
    "Provedor ou modelo",
    "Descrição dos dados",
    "Usa dados pessoais",
    "Nível de risco",
    "Mitigações",
    "Impacto esperado",
    "Status da avaliação",
    "Responsável",
    "Próxima revisão",
    "Política aceita",
    "Aprovação da governança",
    "Usuários estimados",
    "Usuários ativos",
    "Observações",
]
INDICATOR_COLUMNS = [
    "Código da iniciativa",
    "Nome do indicador",
    "Descrição",
    "Unidade",
    "Baseline",
    "Meta",
    "Valor atual",
    "Direção",
    "Responsável",
    "Data de medição",
    "Observações",
]

TEMPLATE_FILES = {
    "initiatives": ("modelo_iniciativas.xlsx", INITIATIVE_COLUMNS),
    "expenses": ("modelo_custos.xlsx", EXPENSE_COLUMNS),
    "ai_cases": ("modelo_casos_ia.xlsx", AI_CASE_COLUMNS),
    "indicators": ("modelo_indicadores.xlsx", INDICATOR_COLUMNS),
}


def create_template(path: Path, columns: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Importação"
    ws.append(columns)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(columns)).coordinate}"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for idx, column in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = max(14, len(column) + 2)
    wb.save(path)
    return path


def ensure_templates(directory: Path) -> dict[str, Path]:
    paths: dict[str, Path] = {}
    for kind, (filename, columns) in TEMPLATE_FILES.items():
        target = directory / filename
        if not target.exists():
            create_template(target, columns)
        paths[kind] = target
    return paths
