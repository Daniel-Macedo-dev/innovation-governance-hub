from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from innovation_governance_hub.persistence.models import (
    ActionItem,
    AIUseCase,
    Expense,
    Initiative,
    InitiativeIndicator,
)
from innovation_governance_hub.services.budget_service import BudgetService
from innovation_governance_hub.services.executive_committee_service import CommitteeBrief
from innovation_governance_hub.services.ui_query_services import PortfolioPriorityRow


def executive_workbook(session: Session) -> bytes:
    output = BytesIO()
    initiatives = session.scalars(select(Initiative)).all()
    ai = session.scalars(select(AIUseCase)).all()
    expenses = session.scalars(select(Expense)).all()
    actions = session.scalars(select(ActionItem)).all()
    summary = pd.DataFrame(
        [
            {"Indicador": "Iniciativas", "Valor": len(initiatives)},
            {"Indicador": "Casos de IA", "Valor": len(ai)},
            {"Indicador": "Aviso", "Valor": "Dados totalmente fictícios"},
        ]
    )
    frames = {
        "Resumo executivo": summary,
        "Iniciativas": pd.DataFrame(
            [
                {
                    "Código": i.code,
                    "Nome": i.name,
                    "Status": i.status,
                    "Estágio": i.current_stage,
                    "Planejado": i.planned_cost,
                    "Realizado": BudgetService(session).initiative_actual(i.id),
                }
                for i in initiatives
            ]
        ),
        "Governança de IA": pd.DataFrame(
            [
                {
                    "Código": x.code,
                    "Nome": x.name,
                    "Risco": x.risk_level,
                    "Status": x.evaluation_status,
                }
                for x in ai
            ]
        ),
        "Custos": pd.DataFrame(
            [
                {
                    "Data": x.competence_date,
                    "Categoria": x.category,
                    "Descrição": x.description,
                    "Status": x.financial_status,
                    "Valor": x.amount,
                }
                for x in expenses
            ]
        ),
        "Pendências": pd.DataFrame(
            [
                {
                    "Descrição": x.description,
                    "Responsável": x.owner,
                    "Prazo": x.deadline,
                    "Status": x.status,
                }
                for x in actions
            ]
        ),
        "Definições das métricas": pd.DataFrame(
            [
                {"Métrica": "Projeto ativo", "Definição": "Não concluído nem arquivado"},
                {"Métrica": "Custo realizado", "Definição": "Soma de despesas Realizado"},
            ]
        ),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in frames.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column in ws.columns:
                ws.column_dimensions[column[0].column_letter].width = min(
                    45, max(12, max(len(str(c.value or "")) for c in column) + 2)
                )
    return output.getvalue()


def committee_workbook(
    session: Session, brief: CommitteeBrief, priorities: list[PortfolioPriorityRow]
) -> bytes:
    output = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    notice = "Dados totalmente fictícios — apoio demonstrativo; decisões permanecem humanas."
    sheets: dict[str, list[list[Any]]] = {
        "Resumo executivo": [
            ["Data da posição", brief.position_date],
            ["Aviso", notice],
            ["Resumo", brief.narrative],
            ["Iniciativas ativas", brief.active_initiatives],
            ["Decisões requeridas", len(brief.decisions)],
        ],
        "Decisões requeridas": [
            [
                "Tipo",
                "Entidade",
                "Motivo",
                "Severidade",
                "Responsável",
                "Prazo",
                "Ação recomendada",
            ],
            *[
                [
                    item.kind,
                    item.entity,
                    item.reason,
                    item.severity,
                    item.owner,
                    item.deadline,
                    item.recommendation,
                ]
                for item in brief.decisions
            ],
        ],
        "Saúde do portfólio": [
            ["Iniciativa", "Saúde", "Motivos"],
            *[[item.name, item.status, "; ".join(item.reasons)] for item in brief.health],
        ],
        "Priorização": [
            [
                "Posição",
                "Código",
                "Iniciativa",
                "Tema",
                "Score",
                "Valor",
                "Esforço",
                "Saúde",
                "Estágio",
                "Responsável",
            ],
            *[
                [
                    index,
                    item.code,
                    item.name,
                    item.theme,
                    item.score,
                    item.value,
                    item.effort,
                    item.health,
                    item.stage,
                    item.owner,
                ]
                for index, item in enumerate(priorities, 1)
            ],
        ],
        "Indicadores": [
            [
                "Iniciativa",
                "Indicador",
                "Unidade",
                "Baseline",
                "Atual",
                "Meta",
                "Direção",
                "Responsável",
                "Medição",
            ],
            *[
                [
                    item.initiative_id,
                    item.name,
                    item.unit,
                    item.baseline_value,
                    item.current_value,
                    item.target_value,
                    item.direction,
                    item.owner,
                    item.measurement_date,
                ]
                for item in session.scalars(select(InitiativeIndicator)).all()
            ],
        ],
        "Governança de IA": [
            ["Métrica", "Valor"],
            *[[item.label, item.value] for item in brief.ai],
        ],
        "Orçamento": [
            ["Métrica", "Valor"],
            *[[item.label, item.value] for item in brief.financial],
        ],
        "Pendências": [
            ["Descrição", "Responsável", "Prazo", "Situação"],
            *[[item.entity, item.owner, item.deadline, item.reason] for item in brief.next_actions],
        ],
        "Mudanças recentes": [["Evento"], *[[item] for item in brief.changes]],
        "Definições": [
            ["Termo", "Definição"],
            ["Score", "70% valor normalizado e 30% facilidade de execução"],
            ["Saúde", "Classificação explicável por prazo, bloqueio, pendência e atividade"],
            ["Aviso", notice],
        ],
    }
    header_fill = PatternFill("solid", fgColor="17365D")
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for column in range(1, sheet.max_column + 1):
            values = [
                len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)
            ]
            sheet.column_dimensions[get_column_letter(column)].width = min(
                55, max(12, max(values, default=10) + 2)
            )
        sheet.sheet_view.showGridLines = False
    workbook.save(output)
    return output.getvalue()
