from io import BytesIO

from openpyxl import load_workbook

from innovation_governance_hub.excel.exporters import committee_workbook
from innovation_governance_hub.services.executive_committee_service import CommitteeBrief


def test_committee_workbook_has_ten_formatted_sheets(session):
    brief = CommitteeBrief(
        position_date="27/07/2026",
        active_initiatives=8,
        decisions=(),
        health=(),
        financial=(),
        ai=(),
        indicators=(),
        next_actions=(),
        changes=(),
        narrative="Cenário executivo fictício.",
    )

    payload = committee_workbook(session, brief, [])
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == [
        "Resumo executivo",
        "Decisões requeridas",
        "Saúde do portfólio",
        "Priorização",
        "Indicadores",
        "Governança de IA",
        "Orçamento",
        "Pendências",
        "Mudanças recentes",
        "Definições",
    ]
    assert workbook["Resumo executivo"]["B2"].value == "27/07/2026"
    assert "fictícios" in workbook["Resumo executivo"]["B4"].value
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref
        assert sheet["A1"].fill.fgColor.rgb.endswith("17365D")
