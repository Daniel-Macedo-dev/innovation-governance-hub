from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api import app
from innovation_governance_hub.excel.exporters import executive_workbook


def test_workbook_has_required_sheets(session):
    wb = load_workbook(BytesIO(executive_workbook(session)))
    assert {
        "Resumo executivo",
        "Iniciativas",
        "Governança de IA",
        "Custos",
        "Pendências",
        "Definições das métricas",
    } <= set(wb.sheetnames)


def test_health():
    client = TestClient(app)
    assert client.get("/health").status_code == 200
    assert client.post("/api/v1/automations/run", json={}).status_code == 401
    summary = client.get("/api/v1/automations/weekly-summary")
    assert summary.status_code == 200
    assert "situacao_orcamentaria" in summary.json()


def test_callback_requires_token_and_reports_missing_notification():
    body = {"fingerprint": "inexistente", "delivery_status": "Enviada"}
    with TestClient(app) as client:
        assert client.post("/api/v1/notifications/callback", json=body).status_code == 401
        response = client.post(
            "/api/v1/notifications/callback",
            json=body,
            headers={"Authorization": "Bearer change-me-local"},
        )
    assert response.status_code == 404
